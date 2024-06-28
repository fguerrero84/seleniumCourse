import openpyxl



class angularPracticeData:

    test_angularPracticeData = [{'first_name':'fer', 'lastname':'fern@test.com', 'gender':'Male'}, {'first_name':'zoe', 'email':'zoe@test.com', 'gender':'female'}]
    @staticmethod
    def excelExtraction(test_case_name):
        book = openpyxl.load_workbook('C:\\Users\\fguerrero\\Downloads\\pythonDemo.xlsx')
        sheet = book.active
        dict = {}
        cell = sheet.cell(row=1, column=2)
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # this is used when you want specficic test case data
                for c in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=i, column=c).value

        return [dict]