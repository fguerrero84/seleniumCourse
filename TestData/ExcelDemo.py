import openpyxl

book = openpyxl.load_workbook('C:\\Users\\fguerrero\\Downloads\\pythonDemo.xlsx')
sheet = book.active
dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = ('Fern')
print(sheet.cell(row=2, column=2).value)
print(sheet['B2'].value)

print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == 'TestCase5': # this is used when you want specficic test case data
        for c in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=i, column=c).value


print(dict)
